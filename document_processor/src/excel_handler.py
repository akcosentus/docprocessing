"""Excel output handling for writing extraction results to an Excel workbook."""

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from filelock import FileLock, Timeout
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.utils.exceptions import InvalidFileException

from src.logger import get_logger

logger = get_logger(__name__)


class WorkbookLockedError(Exception):
    """Raised when the target Excel file is locked or open by another process."""


# Column headers in required order
HEADERS: List[str] = [
    "Timestamp",
    "Facility",
    "Input File",
    "Last Name",
    "First Name",
    "Middle Initial",
    "DOB",
    "SSN",
    "Sex",
    "Street",
    "City",
    "State",
    "Zip",
    "Phone",
    "Primary Insurance",
    "Primary Policy #",
    "Secondary Insurance",
    "Secondary Policy #",
    "Rendering Facility",
    "Diagnosis 1",
    "Diagnosis 2",
    "Diagnosis 3",
    "Diagnosis 4",
    "Confidence",
    "Flags",
    "Notes",
]


def _get_nested(d: Dict[str, Any], dotted_key: str) -> Any:
    """Safely retrieve a nested value from a dict using dot-notation keys.

    Args:
        d: Source dictionary.
        dotted_key: Dot-separated key path (e.g. 'patient.address.city').

    Returns:
        The value at the path, or None if any segment is missing.
    """
    current: Any = d
    for part in dotted_key.split("."):
        if isinstance(current, dict) and part in current:
            current = current[part]
        else:
            return None
    return current


def _build_row(result: Dict[str, Any], input_filename: str) -> tuple[List[Any], Dict[str, Any]]:
    """Build a flat row of values from the nested extraction result.

    Args:
        result: The extraction result dictionary (nested schema).
        input_filename: Original input filename.

    Returns:
        A tuple of (list of cell values matching HEADERS order, result dict for flag checking).
    """
    diagnoses = _get_nested(result, "clinical.diagnoses") or []
    # Pad diagnoses to exactly 4 entries
    padded_diagnoses: List[Any] = list(diagnoses[:4]) + [None] * (4 - len(diagnoses[:4]))

    confidence = _get_nested(result, "_meta.confidence")
    flags = _get_nested(result, "_meta.flags") or []
    missing_required = _get_nested(result, "_meta.missing_required") or []
    
    # Build Notes column summarizing flags in plain English
    notes_parts = []
    consistency_conflict_flags = [f for f in flags if f.startswith("consistency_conflict_")]
    if consistency_conflict_flags:
        notes_parts.append("Some fields had conflicting values between two reads")
    if "low_confidence_single_pass" in flags:
        notes_parts.append("Some fields were only found in one of two reads")
    if missing_required:
        notes_parts.append("Some required fields were not found on the document")
    # Add other flags that aren't covered above
    other_flags = [f for f in flags if not f.startswith("consistency_conflict_") and f not in ["low_confidence_single_pass"]]
    if other_flags:
        notes_parts.append(f"Other flags: {', '.join(other_flags)}")
    
    notes = ". ".join(notes_parts) if notes_parts else None

    row = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        _get_nested(result, "_meta.facility_id"),
        input_filename,
        _get_nested(result, "patient.last_name"),
        _get_nested(result, "patient.first_name"),
        _get_nested(result, "patient.middle_initial"),
        _get_nested(result, "patient.date_of_birth"),
        _get_nested(result, "patient.ssn"),
        _get_nested(result, "patient.sex"),
        _get_nested(result, "patient.address.street"),
        _get_nested(result, "patient.address.city"),
        _get_nested(result, "patient.address.state"),
        _get_nested(result, "patient.address.zip"),
        _get_nested(result, "patient.phone"),
        _get_nested(result, "insurance.primary.insurance_name"),
        _get_nested(result, "insurance.primary.policy_number"),
        _get_nested(result, "insurance.secondary.insurance_name"),
        _get_nested(result, "insurance.secondary.policy_number"),
        _get_nested(result, "clinical.rendering_facility"),
        padded_diagnoses[0],
        padded_diagnoses[1],
        padded_diagnoses[2],
        padded_diagnoses[3],
        confidence,
        ", ".join(flags) if flags else None,
        notes,
    ]
    
    return row, result


def _get_field_column_mapping() -> Dict[str, int]:
    """Return mapping of field paths to column indices (0-based).
    
    Returns:
        Dictionary mapping field paths (e.g., 'patient.last_name') to column indices.
    """
    return {
        "patient.last_name": 3,
        "patient.first_name": 4,
        "patient.middle_initial": 5,
        "patient.date_of_birth": 6,
        "patient.ssn": 7,
        "patient.sex": 8,
        "patient.address.street": 9,
        "patient.address.city": 10,
        "patient.address.state": 11,
        "patient.address.zip": 12,
        "patient.phone": 13,
        "insurance.primary.insurance_name": 14,
        "insurance.primary.policy_number": 15,
        "insurance.secondary.insurance_name": 16,
        "insurance.secondary.policy_number": 17,
        "clinical.rendering_facility": 18,
    }


def _apply_cell_formatting(
    ws, row_num: int, result: Dict[str, Any], row_data: List[Any]
) -> None:
    """Apply color coding and comments to cells based on confidence and flags.
    
    Args:
        ws: The worksheet to format.
        row_num: The row number (1-based).
        result: The extraction result dictionary.
        row_data: The row data list.
    """
    
    # Color Confidence column based on value
    confidence_col = HEADERS.index("Confidence") + 1  # 1-based
    confidence_cell = ws.cell(row=row_num, column=confidence_col)
    confidence_value = row_data[confidence_col - 1]
    
    if confidence_value == "HIGH":
        confidence_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif confidence_value == "MEDIUM":
        confidence_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    elif confidence_value == "LOW":
        confidence_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Highlight flagged field cells
    flags = _get_nested(result, "_meta.flags") or []
    missing_required = _get_nested(result, "_meta.missing_required") or []
    
    field_mapping = _get_field_column_mapping()
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Check for consistency_conflict flags (format: consistency_conflict_<field_path with dots>)
    for flag in flags:
        if flag.startswith("consistency_conflict_"):
            field_path = flag.replace("consistency_conflict_", "")
            if field_path in field_mapping:
                col_idx = field_mapping[field_path] + 1  # 1-based
                cell = ws.cell(row=row_num, column=col_idx)
                if cell.value is None:
                    cell.fill = yellow_fill
                    cell.comment = Comment(
                        "Two different values were detected on two reads — please verify against the original document",
                        "System"
                    )
    
    # Check for low_confidence_single_pass flag (applies to all one-sided fields)
    # Since we don't know which specific fields, we'll check all null fields in the row
    # But only if the flag exists - we'll highlight null fields that are in missing_required
    # Actually, low_confidence_single_pass is a general flag, not per-field
    # We need to check which fields are actually null due to this
    # For now, we'll only highlight if there's a specific field flag or if it's in missing_required
    
    # Check for missing_required fields
    for field_path in missing_required:
        if field_path in field_mapping:
            col_idx = field_mapping[field_path] + 1  # 1-based
            cell = ws.cell(row=row_num, column=col_idx)
            if cell.value is None:
                # Only add comment if not already set
                if cell.comment is None:
                    cell.fill = yellow_fill
                    cell.comment = Comment(
                        "This required field was not found on the document",
                        "System"
                    )
    
    # For low_confidence_single_pass, this is a general flag indicating some fields
    # were found in only one pass. Since one-sided fields keep their value (not null),
    # we can't highlight them by checking for null. The flag is noted in the Notes column.
    # If we need to highlight specific one-sided fields, we'd need to track them separately,
    # but for now the Notes column covers this case.


def append_to_workbook(
    result: Dict[str, Any], input_filename: str, workbook_path: str
) -> Path:
    """Append an extraction result as a new row in an Excel workbook.

    If the workbook does not exist it is created with a header row.
    All records are written to the main "Extractions" sheet with visual
    formatting (color-coded confidence, highlighted flagged fields).

    Args:
        result: The extraction result dictionary (nested schema).
        input_filename: Original input filename (without directory path).
        workbook_path: Path to the .xlsx file to create or append to.

    Returns:
        Path to the workbook file.

    Raises:
        WorkbookLockedError: If the file cannot be opened because it is
            locked or in use by another process.
    """
    logger.debug(f"Appending to workbook: {workbook_path}")
    wb_path = Path(workbook_path)
    wb_path.parent.mkdir(parents=True, exist_ok=True)

    row_data, result_dict = _build_row(result, input_filename)

    lock_path = f"{workbook_path}.lock"
    lock = FileLock(lock_path, timeout=30)

    try:
        with lock:
            if wb_path.exists():
                logger.debug(f"Opening existing workbook: {wb_path}")
                try:
                    wb = load_workbook(wb_path)
                except PermissionError as exc:
                    raise WorkbookLockedError(
                        f"Cannot open '{wb_path}' — the file is locked or open in another "
                        "application. Close it and try again."
                    ) from exc
                except InvalidFileException as exc:
                    raise WorkbookLockedError(
                        f"Cannot open '{wb_path}' — the file appears corrupted or is not a "
                        "valid .xlsx file."
                    ) from exc

                # Get or create the main sheet
                if "Extractions" in wb.sheetnames:
                    ws = wb["Extractions"]
                else:
                    ws = wb.active
                    ws.title = "Extractions"
                    ws.append(HEADERS)
            else:
                logger.debug(f"Creating new workbook: {wb_path}")
                wb = Workbook()
                ws = wb.active
                ws.title = "Extractions"
                ws.append(HEADERS)

            # Append data row
            ws.append(row_data)
            row_num = ws.max_row
            
            # Apply formatting (color coding, comments)
            _apply_cell_formatting(ws, row_num, result_dict, row_data)
            
            # Auto-size all columns
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[col_letter].width = adjusted_width
            
            logger.debug("Appended row to Extractions sheet")

            try:
                wb.save(wb_path)
                logger.debug(f"Saved workbook: {wb_path}")
            except PermissionError as exc:
                raise WorkbookLockedError(
                    f"Cannot save '{wb_path}' — the file is locked or open in another "
                    "application. Close it and try again."
                ) from exc

    except Timeout as exc:
        raise WorkbookLockedError(
            f"Cannot acquire lock on '{workbook_path}' after 30 seconds — "
            "the file may be open in Excel or another process holds the lock."
        ) from exc

    return wb_path

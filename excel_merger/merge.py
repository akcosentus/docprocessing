#!/usr/bin/env python3
"""Excel file merger - combines multiple Excel files into one master file."""

import argparse
import sys
from pathlib import Path
from typing import List, Set

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Expected headers (normalized: lowercase, stripped)
EXPECTED_HEADERS = {
    "patient number",
    "sms",
    "practicename",
    "patientid",
    "patientdob",
    "patientfirstname",
    "patientlastname",
    "patientbalance",
    "patientcreditpresent",
    "1ststatementdate",
    "laststatementdate",
    "lastoutreachdate",
    "lastpaymentdate",
    "balanceage",
    "aginggroup",
    "statementcount",
    "smscount",
    "emailcount",
    "callcount",
    "autopay",
    "paymentplan",
    "endofcadence",
    "address",
    "addressline2",
    "city",
    "state",
    "zip",
    "email",
    "servicedate",
    "physicianname",
    "primaryinsurance",
}


def normalize_header(header: str) -> str:
    """Normalize a header string: strip whitespace and lowercase."""
    return header.strip().lower()


def validate_headers(file_path: Path) -> tuple[bool, List[str], List[str]]:
    """
    Validate that a file's headers match the expected headers.
    
    Args:
        file_path: Path to the Excel file to validate.
        
    Returns:
        Tuple of (is_valid, missing_columns, extra_columns)
    """
    try:
        df = pd.read_excel(file_path, nrows=0)  # Read only headers
        file_headers = {normalize_header(h) for h in df.columns}
        expected_set = EXPECTED_HEADERS
        
        missing = sorted(expected_set - file_headers)
        extra = sorted(file_headers - expected_set)
        
        is_valid = len(missing) == 0 and len(extra) == 0
        return is_valid, missing, extra
    except Exception as e:
        print(f"❌ Error reading {file_path.name}: {e}", file=sys.stderr)
        return False, [], []


def merge_excel_files(input_dir: str, output_path: str) -> None:
    """
    Merge multiple Excel files into one master file.
    
    Args:
        input_dir: Directory containing Excel files to merge.
        output_path: Path to write the merged output file.
    """
    input_path = Path(input_dir)
    if not input_path.exists() or not input_path.is_dir():
        print(f"❌ Error: Input directory does not exist: {input_dir}", file=sys.stderr)
        sys.exit(1)
    
    # Find all .xlsx files
    excel_files = sorted(input_path.glob("*.xlsx"))
    if not excel_files:
        print(f"❌ Error: No .xlsx files found in {input_dir}", file=sys.stderr)
        sys.exit(1)
    
    # Validate and collect valid files
    valid_files = []
    skipped_files = []
    
    for file_path in excel_files:
        is_valid, missing, extra = validate_headers(file_path)
        if is_valid:
            valid_files.append(file_path)
        else:
            reason_parts = []
            if missing:
                reason_parts.append(f"missing columns: {missing}")
            if extra:
                reason_parts.append(f"extra columns: {extra}")
            reason = " — ".join(reason_parts)
            skipped_files.append((file_path, reason))
    
    if not valid_files:
        print("❌ Error: No valid files to merge (all files failed validation)", file=sys.stderr)
        sys.exit(1)
    
    # Merge valid files
    merged_dataframes = []
    merge_report = []
    
    for file_path in valid_files:
        try:
            df = pd.read_excel(file_path)
            # Add source_file column as first column
            df.insert(0, "source_file", file_path.name)
            merged_dataframes.append(df)
            merge_report.append((file_path.name, len(df)))
        except Exception as e:
            print(f"❌ Error reading {file_path.name}: {e}", file=sys.stderr)
            skipped_files.append((file_path, f"read error: {e}"))
    
    # Combine all dataframes
    if not merged_dataframes:
        print("❌ Error: No data to merge", file=sys.stderr)
        sys.exit(1)
    
    master_df = pd.concat(merged_dataframes, ignore_index=True)
    total_rows = len(master_df)
    
    # Write output using openpyxl for formatting
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    # Write to Excel with openpyxl formatting
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        master_df.to_excel(writer, index=False, sheet_name="Sheet1")
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets["Sheet1"]
        
        # Format header row: bold, Arial font
        header_font = Font(bold=True, name="Arial")
        for cell in worksheet[1]:
            cell.font = header_font
        
        # Auto-size column widths
        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            worksheet.column_dimensions[col_letter].width = adjusted_width
        
        # Freeze top row
        worksheet.freeze_panes = "A2"
        
        # Set Arial font for all cells
        arial_font = Font(name="Arial")
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = arial_font
    
    # Print report
    print(f"Confirmed output path: {output_path}")
    print()
    
    for filename, row_count in merge_report:
        print(f"✅ Merged:   {filename} ({row_count} rows)")
    
    for file_path, reason in skipped_files:
        print(f"❌ Skipped:  {file_path.name} — {reason}")
    
    print()
    print(f"Total rows merged: {total_rows}")
    print(f"Output written to: {output_path}")


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Merge multiple Excel files into one master file"
    )
    parser.add_argument(
        "--input",
        required=True,
        help="Input directory containing Excel files to merge",
    )
    parser.add_argument(
        "--output",
        help="Output file path (default: OneDrive path)",
        default=None,
    )
    
    args = parser.parse_args()
    
    # Default output path (OneDrive)
    if args.output is None:
        default_output = Path.home() / "Library/CloudStorage/OneDrive-Cosentus,LLC/Cindy 2026/Cindy Batch Final/Master.xlsx"
        output_path = str(default_output)
    else:
        output_path = args.output
    
    merge_excel_files(args.input, output_path)


if __name__ == "__main__":
    main()

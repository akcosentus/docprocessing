#!/usr/bin/env python3
"""Reconciler - matches a subsheet (phone list) to a master table by phone number."""

import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# One file per folder; pandas loads by extension
SUPPORTED_INPUT_EXTENSIONS = {".csv", ".xlsx", ".xls"}


def to_string(phone) -> str:
    """Convert phone number to string, handling NaN and float formatting."""
    if pd.isna(phone):
        return ""
    phone_str = str(phone)
    if phone_str.endswith(".0"):
        phone_str = phone_str[:-2]
    return phone_str


def normalize_phone_key(phone) -> str:
    """String for matching: strip whitespace, drop leading + (E.164 vs plain digits)."""
    s = to_string(phone).strip()
    if s.startswith("+"):
        s = s[1:]
    return s


def read_table(path: Path) -> pd.DataFrame:
    """Load a tabular file with pandas (CSV or Excel)."""
    suffix = path.suffix.lower()
    try:
        if suffix == ".csv":
            return pd.read_csv(path)
        if suffix in (".xlsx", ".xls"):
            return pd.read_excel(path)
    except Exception as e:
        print(f"Error reading {path.name}: {e}", file=sys.stderr)
        sys.exit(1)
    print(
        f"Error: Unsupported file type {suffix!r} for {path}. "
        f"Use one of: {sorted(SUPPORTED_INPUT_EXTENSIONS)}",
        file=sys.stderr,
    )
    sys.exit(1)


def find_single_table_file(directory: Path, folder_name: str) -> Path:
    """
    Find exactly one supported data file (.csv, .xlsx, .xls) in a directory.

    Raises:
        SystemExit: If directory is missing, or 0 / 2+ matching files
    """
    if not directory.exists():
        print(
            f"Error: {folder_name}/ directory does not exist: {directory}",
            file=sys.stderr,
        )
        sys.exit(1)

    if not directory.is_dir():
        print(
            f"Error: {folder_name}/ is not a directory: {directory}",
            file=sys.stderr,
        )
        sys.exit(1)

    candidates = sorted(
        {
            p.resolve()
            for p in directory.iterdir()
            if p.is_file() and p.suffix.lower() in SUPPORTED_INPUT_EXTENSIONS
        }
    )

    if len(candidates) == 0:
        print(
            f"Error: {folder_name}/ must contain exactly 1 data file "
            f"({', '.join(sorted(SUPPORTED_INPUT_EXTENSIONS))}), found 0",
            file=sys.stderr,
        )
        sys.exit(1)

    if len(candidates) > 1:
        names = ", ".join(Path(p).name for p in candidates)
        print(
            f"Error: {folder_name}/ must contain exactly 1 data file, "
            f"found {len(candidates)}: {names}",
            file=sys.stderr,
        )
        sys.exit(1)

    return Path(candidates[0])


def _find_subsheet_phone_column(df: pd.DataFrame):
    """Subsheet phone column: exact 'To', else case-insensitive 'phone number'."""
    for col in df.columns:
        if str(col).strip() == "To":
            return col
    for col in df.columns:
        if str(col).strip().lower() == "phone number":
            return col
    return None


def reconcile_calls(master_path: str, subsheet_path: str, output_path: str) -> None:
    """
    Match subsheet phone numbers to master rows; output full master rows for each match.

    Args:
        master_path: Master table (.csv / .xlsx / .xls) with 'phone number' column.
        subsheet_path: Subsheet table with 'To' or 'phone number' column.
        output_path: Output .xlsx path.
    """
    master_file = Path(master_path)
    subsheet_file = Path(subsheet_path)

    if not master_file.exists():
        print(f"Error: Master file does not exist: {master_path}", file=sys.stderr)
        sys.exit(1)

    if not subsheet_file.exists():
        print(f"Error: Subsheet file does not exist: {subsheet_path}", file=sys.stderr)
        sys.exit(1)

    master_df = read_table(master_file)

    master_phone_col = None
    for col in master_df.columns:
        if str(col).strip().lower() == "phone number":
            master_phone_col = col
            break

    if master_phone_col is None:
        print("Error: Master file must have a 'phone number' column", file=sys.stderr)
        sys.exit(1)

    subsheet_df = read_table(subsheet_file)

    subsheet_phone_col = _find_subsheet_phone_column(subsheet_df)
    if subsheet_phone_col is None:
        print(
            "Error: Subsheet must have a 'To' or 'phone number' column",
            file=sys.stderr,
        )
        sys.exit(1)

    master_row_count = len(master_df)
    subsheet_row_count = len(subsheet_df)

    subsheet_df = subsheet_df.copy()
    subsheet_df["normalized_sub"] = subsheet_df[subsheet_phone_col].apply(normalize_phone_key)
    subsheet_numbers = set(subsheet_df["normalized_sub"].dropna())
    subsheet_numbers.discard("")

    master_df = master_df.copy()
    master_df["normalized_phone"] = master_df[master_phone_col].apply(normalize_phone_key)

    matched_df = master_df[master_df["normalized_phone"].isin(subsheet_numbers)].copy()
    matched_df = matched_df.drop(columns=["normalized_phone"], errors="ignore")

    matched_row_count = len(matched_df)
    master_phones = set(master_df["normalized_phone"].dropna())
    master_phones.discard("")
    matched_distinct_phones = subsheet_numbers & master_phones
    unmatched_count = len(subsheet_numbers - master_phones)

    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        matched_df.to_excel(writer, index=False, sheet_name="Sheet1")

        workbook = writer.book
        worksheet = writer.sheets["Sheet1"]

        header_font = Font(bold=True, name="Arial")
        for cell in worksheet[1]:
            cell.font = header_font

        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[col_letter].width = adjusted_width

        worksheet.freeze_panes = "A2"

        arial_font = Font(name="Arial")
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = arial_font

    master_filename = Path(master_path).name
    subsheet_filename = Path(subsheet_path).name
    print(f"Master file:        {master_filename} ({master_row_count:,} rows)")
    print(f"Subsheet file:      {subsheet_filename} ({subsheet_row_count:,} rows)")
    print(f"Subsheet phones:    {len(subsheet_numbers):,} distinct non-empty numbers")
    print()
    print(f"Matched phones:     {len(matched_distinct_phones):,} found in master")
    print(f"Unmatched phones:   {unmatched_count:,} not found in master")
    print(
        f"Output rows:        {matched_row_count:,} "
        "(full master rows; duplicates if master has dup phones)"
    )
    print()
    print(f"Output written to: {output_path}")


def main():
    """Main entry point."""
    base_dir = Path.home() / "Desktop" / "processing"
    master_dir = base_dir / "master"
    subsheet_dir = base_dir / "subsheet"

    master_file = find_single_table_file(master_dir, "master")
    subsheet_file = find_single_table_file(subsheet_dir, "subsheet")

    today = datetime.now().strftime("%Y-%m-%d")
    output_filename = f"Matched_To_Master_{today}.xlsx"
    output_path = str(base_dir / output_filename)

    reconcile_calls(str(master_file), str(subsheet_file), output_path)


if __name__ == "__main__":
    main()
